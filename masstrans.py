
from selenium import webdriver
import openpyxl
from chinesewords import zhtoeng

path = r"C:\Users\Richd\Documents\ankisound (4).xlsx"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

start = 1

while sheet_obj.cell(row=start,column=1).value:
    hanzi = sheet_obj.cell(row=start, column=1).value
    pinyin = sheet_obj.cell(row=start, column=2)
    define = sheet_obj.cell(row=start, column=3)
    sound = sheet_obj.cell(row=start, column=4)
    pinyin.value, define.value, sound.value = zhtoeng(hanzi)
    ankisounds = f"[sound:pronunciation_zh_{sound.value}.mp3]"
    sound.value = ankisounds
    start += 1

wb_obj.save(path)


